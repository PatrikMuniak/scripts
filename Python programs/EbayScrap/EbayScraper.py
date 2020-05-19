#!/usr/bin/env Python
import requests
import bs4
import re
import openpyxl
import sys
ebayserp= requests.get('https://www.ebay.co.uk/sch/i.html?_nkw='+
                       '+'.join(sys.argv[1:]), 'rb')

bs=bs4.BeautifulSoup(ebayserp.text, 'html.parser')


#gathering links of ads
links=[]
for link in bs.find_all("h3", class_=re.compile("lvtitle")):
    if link.previous_element.previous_element!='SPONSORED':
        links.append(link.a.get("href"))

#open the links in links(list). Parse with bs. Get title, price and description
data=[('tile','price', 'description')]
for link in links:
    adv_page=requests.get(link)
    adv_soup=bs4.BeautifulSoup(adv_page.text, 'html.parser')
    
    #Title
    title=adv_soup.find("title")
    title=str(title.string)
    title=title.replace('  | eBay','')
    if title == None:
        print('----')
    else:
        print(title)
        
    #Price
    price=adv_soup.find("span", class_=re.compile("notranslate"), id=True,
                        itemprop=True, content=True)
    price=str(price.string)
    if price == None:
        print('----')
    else:
        print(price)

    #Description
    description=adv_soup.find("meta", property=re.compile("og:description"))
    description=str(description.get("content"))
    if title == None:
        print('----')
    else:
        print(description)
    adv_data=(title, price, description)        
    data.append(adv_data)                       

#Creating an Excel workbook with the delattr    
wb = openpyxl.Workbook()   
ws = wb.active
ws.title=str('_'.join(sys.argv[1:]))

r = 1
for tuple_ in data:
    ws.cell(row = r , column = 1 , value = tuple_[0]) #title
    ws.cell(row = r , column = 2 , value = tuple_[1]) #price
    ws.cell(row = r , column = 3 , value = tuple_[2]) #description
    r+=1
os.chdir( r'C:\Users\Patrick\Desktop\\' )
wb.save(str(' '.join(sys.argv[1:])+'.xlsx'))



            
